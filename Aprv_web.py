import streamlit as st  #pip install streamlit
import pandas  as pd
import os
import datetime
import calendar
import sqlite3
import numpy as np
import math
import subprocess

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side

import getpass

master_file_path = '決裁承認.db'
menber_file_path = 'menber.db'
#01 desktop_path = os.getenv("HOMEDRIVE") + os.getenv("HOMEPATH") + "\\Desktop"
desktop_path = ''

lst_Section = ['全部署']
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
#【関数定義】
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
def UserInfo():
    ps = getpass.getuser()
    return ps

def NpRemove(d_str):
    d = str(d_str).replace('[','')
    d = str(d).replace(']','')
    d = str(d).replace("'","")
    d = str(d).replace(" ","")
    return d

def DataMake_Checklist(df,file,dys,dye,sect,path):
    dys = datetime.datetime.date(dys)
    dye = datetime.datetime.date(dye)
    df_Develop = df.copy()
    savefile = path+'//' + file + '_' + str(dye.month) + '月' +'.xlsx'
    df_Develop = df_Develop.groupby(['申請者所属組織','取引先','取引先コード','開発コード'], as_index = False).sum(numeric_only = True)
    df_Develop = df_Develop[['取引先コード','取引先','合計(税抜)','合計(税込)']]
    df_Develop.loc['合計','合計(税抜)'] = df_Develop['合計(税抜)'].sum(numeric_only=True)
    df_Develop.loc['合計','合計(税込)'] = df_Develop['合計(税込)'].sum(numeric_only=True)
    df_Develop = df_Develop.sort_values('取引先コード')
    df_Develop['合計(税抜)'] = df_Develop['合計(税抜)'].astype("int64")
    df_Develop['合計(税込)'] = df_Develop['合計(税込)'].astype("int64")
    df_Develop.style.format('{:,d}')
    df_Develop.to_excel(savefile, encoding="cp932", index=False)
    #------------------ エクセル操作 ---------------
    wb = openpyxl.load_workbook(savefile)
    ws = wb['Sheet1']
    ws.insert_rows(1)
    ws['A1'] = str(dys)+'～'+str(dye) + '【' + str(sect) + '】' + file
    ws['B'+str(len(df_Develop)+2)] = '合計'
    ws['A1'].font = Font(color = '00ff00ff',size = 15 ,italic = False , bold = True)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    for num in range(len(df_Develop)):
        ws.cell(num+3,3).number_format = '#,##0'
        ws.cell(num+3,4).number_format = '#,##0'
        ws.cell(num+3,1).alignment = Alignment(horizontal = 'center', 
                                    vertical = 'center',
                                    wrap_text = False)
    # 罫線(外枠)を設定
    border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'), 
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    # セルに罫線を設定
    for row_num in range(2,3+len(df_Develop)):
        for col_num in range(1,5):
            ws.cell(row=row_num ,column=col_num).border = border

    wb.save(savefile)
    wb.close()
    #------------------ エクセルファイルオープン ---------------
    subprocess.Popen(['start',savefile],shell=True)

def DataMake_Agg(df,file,path,dye):
    df_Stock = df.copy()
    dye = datetime.datetime.date(dye)
    savefile = path+'//' + file + '_' + str(dye.month) + '月' +'.csv'
    df_SectCode = pd.read_csv ('部署コード.csv')
    df_Stock['申請者所属組織'] = df_Stock['申請者所属組織'].replace(df_SectCode['部署'].to_list(),df_SectCode['ID'].to_list())
    df_Stock['開発orNot'] = np.where(df_Stock['開発orNot'] == 'Not開発', '', '1')
    df_Stock['開発コード'] = np.where(df_Stock['開発コード'] == '999', '', '1')
    # 0埋め・・・注意点：欠損があるとfloatになる➡小数点が出る　文字列型にしないとzfillは使えない
    #df_Stock['取引先コード'] = df_Stock['取引先コード'].where(df_Stock['取引先コード'].notnull(),0)
    df_Stock['取引先コード'] = df_Stock['取引先コード'].astype("int64")
    df_Stock['取引先コード'] = df_Stock['取引先コード'].astype("str")
    df_Stock['取引先コード'] = df_Stock['取引先コード'].str.zfill(7)
    # 取引先で合算したいが、文字列や数値をプロテクトしたいものがあるので複数列をキーとする
    df_Stock = df_Stock.groupby(['請求確定用(請求・分納)',
                                    '申請者所属組織',
                                    '開発orNot',
                                    '取引先',
                                    '取引先コード',
                                    '開発コード'
                                    ], 
                                    as_index = False).sum(numeric_only = True)
    df_Stock = df_Stock[['請求確定用(請求・分納)',
                         '申請者所属組織',
                        '開発orNot',
                        '取引先コード',
                        '開発コード',
                        '合計(税込)'
                        ]]
    df_Stock = df_Stock.sort_values('取引先コード')
    df_Stock.insert(4,'消費税課税区分',2)
    df_Stock.insert(5,'品名コード',99999)
    df_Stock.insert(8,'備考','')
    df_Stock = df_Stock.set_axis(['仕入日',
                    '仕入れ部門コード',
                    '研究開発区分',
                    '手配先コード',
                    '消費税課税区分',
                    '品名コード',
                    '開発コード',
                    '仕入金額',
                    '備考'],
                        axis=1)
    df_Stock.to_csv(savefile,
                    date_format='%Y/%m/%d',
                        encoding="cp932",
                        index=False)

def DataMake_Item(df,file,path,dye):
    df_Item = df.copy()
    dye = datetime.datetime.date(dye)
    savefile = path+'//' + file + '_' + str(dye.month) + '月' +'.csv'
    df_Item = df_Item.sort_values('取引先コード')
    lst_15P = ['数量','単価','金額','商品名','分類']
    df_NewItem = pd.DataFrame(columns = df_Item.columns)
    for num_1 in df_Item.index:
        cnt_m = 0
        for num_2 in range(1,16):
            if df_Item.loc[num_1,'数量'+str(num_2)] is not None:
                if math.isnan(df_Item.loc[num_1,'数量'+str(num_2)]):
                    # ダミー処理
                    a=0
                else:
                    cnt_m = cnt_m + 1
                    index =  str(num_1) + '_' + str(num_2)
                    df_NewItem.loc[index] = df_Item.loc[num_1]
                    for num_3 in range(len(lst_15P)):
                        df_NewItem.loc[index,lst_15P[num_3]] = df_NewItem.loc[index,lst_15P[num_3]+str(num_2)]
                    df_NewItem.loc[index,'明細番号'] = cnt_m
                    df_NewItem.loc[index,'申請書番号'] = num_1

    df_NewItem = df_NewItem[['取引先コード',
                            '取引先',
                            '開発orNot',
                            '開発コード',
                            '数量',
                            '単価',
                            '金額',
                            '明細番号',
                            '商品名',
                            '分類',
                            '申請者名',
                            '請求確定用(請求・分納)',
                            '申請書番号',
                            '申請者所属組織']]
    
    df_NewItem = df_NewItem.set_axis(['発注先コード',
                                    '発注先',
                                    '新規開発/NOT開発',
                                    '開発コード',
                                    '数量',
                                    '単価（税抜）',
                                    '金額（税抜）',
                                    '明細番号',
                                    '品名',
                                    '分類',
                                    '申請者',
                                    '取得年月日',
                                    '申請書番号',
                                    '業務機能'],
                                    axis=1)
    df_NewItem.to_csv(savefile, encoding="cp932", index=False)
#◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇
# db　読込・保存
#◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇◇
def SQL_read(file,index):
    conn = sqlite3.connect(file)
    cur = conn.cursor()
    df = pd.read_sql('SELECT * FROM sample', conn,index_col=index)
    cur.close()
    conn.close()
    return df
def SQL_write(df,file):
    conn = sqlite3.connect(file)
    cur = conn.cursor()
    df.to_sql('sample', conn, if_exists='replace',index = True)
    cur.close()
    conn.close()
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
#【コールバック関数】
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
def callback_billset():
    print('----------------------コールバックテスト～請求書情報確定---------------------')
    print('')
    print('---------------------------------------------------------------------')
def callback_billedit():
    print('----------------------コールバックテスト～請求書情報編集---------------------')
    print('')
    print('---------------------------------------------------------------------')
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
#【マスターデータ読み込み】
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
User = UserInfo()
flg_Master = False
if os.path.exists(master_file_path):
    print('-------------read_db-------------')
    df_Master = SQL_read(master_file_path,'自動連番')
    lst_Section = df_Master['申請者所属組織'].unique().tolist()
    lst_Section.append('全部署')
    flg_Master = True
#------- ユーザーデータ管理
if os.path.exists(menber_file_path):
    #既存データ読み込み
    df_Menber = SQL_read(menber_file_path,'ID')
else:
    #データフレーム新規作成
    df_Menber = pd.DataFrame(columns = ['Date','UseInfo_01']) 
    df_Menber.index.name = "ID"
    df_Menber.loc[User,'UseInfo_01'] = '全部署'    
df_Menber.loc[User,'Date'] = datetime.date.today()
print(df_Menber)
SQL_write(df_Menber,menber_file_path)
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
#【ブラウザ画面作成】
# 絞り込み、ファイル読み込み、決裁承認リスト
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
st.set_page_config(page_title='請求書管理',layout='wide')
st.subheader('決裁承認申請書一覧')
TextHorder = st.text('【Info】')
#--------------------------------------------
# サイドバー
#--------------------------------------------
with st.sidebar:
    #---CSV読み込み
    with st.form(key='Append_form'):
        st.subheader('データ更新＠決裁承認担当者')
        file = st.file_uploader('選択したCSVファイルをデータベースに上書きします', type="CSV")
        if st.form_submit_button('データ更新'):
            if flg_Master:
                df_Read = pd.read_csv(file, encoding="cp932",index_col=0)
                if set(df_Read.columns) == set(df_Master.columns):
                    df_Merge = pd.merge(df_Master,df_Read, on=['自動連番'], how='outer', indicator=True)
                    df_Keep = df_Master[df_Merge['_merge'].head(len(df_Master)) == 'left_only']
                    df_Master = pd.concat([df_Keep,df_Read])
                    Cnt_Merge = df_Merge['_merge'].value_counts()
                    Cnf_master_Bf = len(df_Master)
                    #請求書確定となった行を消す（分納は未対応） 引継ぎでA-***** 以外外指定されないか要確認
                    lst_FinNo = (df_Master['申請書No_(請求・分納・発注)'].dropna()).to_list()
                    df_Master = df_Master.drop(lst_FinNo,errors='ignore')
                    Cnt_Del = Cnf_master_Bf - len(df_Master)
                    TextHorder.write('【Info】' 
                                    + str(Cnt_Merge['right_only']) + '件追加 ' 
                                    + str(Cnt_Merge['both']) + '件更新 → 請求書確定により'
                                    + str(Cnt_Del) + '件削除')
                else:
                    TextHorder.write('【Info】'+'読込データが不正です')                    
            else:
                df_Master = pd.read_csv(file, encoding="cp932",index_col=0)
                TextHorder.write('【Info】'+ str(len(df_Master)) + '件のデータを新規登録しました')
            df_Master = df_Master.sort_index()
            lst_Section = df_Master['申請者所属組織'].unique().tolist()
            lst_Section.append('全部署')
            print('-------------write_db-------------')
            SQL_write(df_Master,master_file_path)
            flg_Master = True
    #---絞り込み
    with st.form(key='Refine_foram'):
        print('================== RefineMenu Set')
        st.subheader('ファイル出力')
        #Cnt = len(lst_Section) - 1 
        #box_Section = st.selectbox('部署',lst_Section,index = lst_Section.index(df_Menber.loc[User,'UseInfo_01']))
        #print(df_Menber.loc[User,'UseInfo_01'].split(","))
        #01 box_Section = st.multiselect('部署',lst_Section,default=df_Menber.loc[User,'UseInfo_01'].split(","))
        box_Section = st.multiselect('部署',lst_Section,default=['全部署'])
        dt = datetime.date.today()
        dt = dt - datetime.timedelta(days=10)
        dt_st = dt.replace(day=1)
        dt_end = dt.replace(day=calendar.monthrange(dt.year, dt.month)[1])
        col1,col2 = st.columns(2)
        with col1:
            box_BillStart = st.date_input('請求確定[start]',dt_st)
        with col2:
            box_BillEnd = st.date_input('請求確定[end]',dt_end)        
        btn_Palset = st.form_submit_button('部署と請求日の確定')
        box_Output = st.selectbox('出力形式選択',["請求書リスト","一般仕入（集計）","仕入明細"],index = 0)
        btn_Refine = st.form_submit_button('ファイル出力')
        if btn_Palset:
            #ユーザー情報格納
            print('')
            df_Menber.loc[User,'UseInfo_01'] = NpRemove(box_Section)
            SQL_write(df_Menber,menber_file_path)
        if btn_Refine:
            #ユーザー情報格納
            df_Menber.loc[User,'UseInfo_01'] = NpRemove(box_Section)
            SQL_write(df_Menber,menber_file_path)
            #各ファイル作成
            if flg_Master:
                df_OrignData = df_Master.copy()
                df_OrignData = df_OrignData[df_OrignData.index.str.contains("C", na=False)]
                #新規取引先を取引先にコピー
                df_OrignData['取引先'] = df_OrignData['取引先'].where(df_OrignData['取引先'].notnull(),df_OrignData['新規取引先'])
                # 取引先コード 空欄を0に　DASHは0にするとダメかも・・・
                df_OrignData['取引先コード'] = df_OrignData['取引先コード'].where(df_OrignData['取引先コード'].notnull(),0)
                # 開発コード 空欄を999に　DASHは0にするとダメかも・・・
                df_OrignData['開発コード'] = df_OrignData['開発コード'].where(df_OrignData['開発コード'].notnull(),999)
                #所属絞り込み
                if '全部署' not in box_Section:
                    df_OrignData = df_OrignData[df_OrignData['申請者所属組織'].isin(box_Section) ]
                #----------------期間絞り込み
                #日付型指定 
                df_OrignData["決裁日"] = pd.to_datetime(df_OrignData["決裁日"])
                df_OrignData['請求確定用(請求・分納)'] = pd.to_datetime(df_OrignData['請求確定用(請求・分納)'])
                #期間絞り込み
                time_st = datetime.datetime(box_BillStart.year,box_BillStart.month,box_BillStart.day,00,00,00)
                time_ed = datetime.datetime(box_BillEnd.year,box_BillEnd.month,box_BillEnd.day,23,59,59)
                df_OrignData = df_OrignData[(df_OrignData['請求確定用(請求・分納)'] >= time_st) | 
                                            df_OrignData['請求確定用(請求・分納)'].isnull()]
                df_OrignData = df_OrignData[(df_OrignData['請求確定用(請求・分納)'] <= time_ed) | 
                                            df_OrignData['請求確定用(請求・分納)'].isnull()]
                if box_Output == '請求書リスト':
                    DataMake_Checklist(df_OrignData,box_Output,time_st,time_ed,box_Section,desktop_path)
                if  box_Output == '一般仕入（集計）':
                    DataMake_Agg(df_OrignData,box_Output,desktop_path,time_ed)
                TextHorder.write('【Info】「'+ box_Output + '」ファイルをデスクトップに作成しました')
                if  box_Output == '仕入明細':
                    DataMake_Item(df_OrignData,box_Output,desktop_path,time_ed)
            else:
                TextHorder.write('【Info】'+'決裁承認申請書が登録されていません')

#--------------------------------------------
# メイン
#--------------------------------------------
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
# 決裁承認リスト表示
#◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇◆◇
with st.form(key='form_main'):
    print('--------------------------main run')
    if flg_Master:
        print('--------------------------data reroad')
        df_Dsp = df_Master.copy()
        ##請求書確定となった行を消す（分納は未対応） ➡　マスターから消すことにする
        #lst_FinNo = (df_Dsp['申請書No_(請求・分納・発注)'].dropna()).to_list()
        #df_Dsp = df_Dsp.drop(lst_FinNo,errors='ignore')

        df_Dsp = df_Dsp[['申請日',
                            '申請者所属組織',
                            '申請者名',
                            '申請事項',
                            '取引先',
                            'アクセスURL',
                            '合計(税込)',
                            '請求確定用(請求・分納)']].copy()
        df_Dsp['請求確定用(請求・分納)'] = pd.to_datetime(df_Dsp['請求確定用(請求・分納)'])
        #---------------絞り込み適用
        col1,col2,col3 = st.columns(3)
        with col1:
            btn_section = st.checkbox('部署指定を有効にする',value=True)
        with col2:
            btn_Billon = st.checkbox('請求書確定日を有効にする',value=True)
        with col3:
            btn_bill = st.checkbox('請求書未処理を表示しない',value=False)
        if btn_section:
            if '全部署' not in box_Section:
                df_Dsp = df_Dsp[df_Dsp['申請者所属組織'].isin(box_Section)]
        if btn_Billon:
            time_st = datetime.datetime(box_BillStart.year,box_BillStart.month,box_BillStart.day,00,00,00)
            time_ed = datetime.datetime(box_BillEnd.year,box_BillEnd.month,box_BillEnd.day,23,59,59)
            df_Dsp = df_Dsp[(df_Dsp['請求確定用(請求・分納)'] >= time_st) | df_Dsp['請求確定用(請求・分納)'].isnull()]
            df_Dsp = df_Dsp[(df_Dsp['請求確定用(請求・分納)'] <= time_ed) | df_Dsp['請求確定用(請求・分納)'].isnull()]
        if btn_bill:
            #----------None判定
            df_Dsp = df_Dsp[df_Dsp['請求確定用(請求・分納)'].isnull() == False]
        btn_Fin = st.form_submit_button('チェックボックス反映',on_click = callback_billset)
        #st.subheader('決裁承認申請書リスト')
        #st.dataframe(df_Dsp)
        st.dataframe(
            df_Dsp,
            column_config={
                'アクセスURL': st.column_config.LinkColumn(
                    # 表示するカラム名
                    "Link",
                    # 表示データのテキスト
                    #display_text = '自動連番'
                )
            },
        )
        if btn_Fin:
            TextHorder.write('【Info】部署／請求書確定日の絞り込みは「ファイル出力」か「条件確定」をしないと反映されないのでご注意ください')
        st.text(str(len(df_Dsp)) + ' Data')
        st.text('金額合計：¥' + str(format(df_Dsp['合計(税込)'].sum(),',')) + '（税込）')
    else:
        TextHorder.write('【Main_Info】データの登録がありません')
